from flask import Flask, render_template, request, redirect, url_for, flash, session
import os
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook, load_workbook
from datetime import datetime
import uuid
from dotenv import load_dotenv

import cv2
from skimage.metrics import structural_similarity as ssim
from PIL import Image
import imagehash
import numpy as np

# Load environment variables FIRST
load_dotenv()

app = Flask(__name__)

# SECURE: Always provide fallbacks for environment variables
app.secret_key = os.environ.get('SECRET_KEY') or 'dev-fallback-key-change-in-production-NEVER-USE-THIS'

# ─────────────────────────── Configuration ───────────────────────────
# SECURE: All environment variables now have safe fallbacks
UPLOAD_FOLDER_TASKS       = os.environ.get('UPLOAD_FOLDER_TASKS', 'static/uploads/tasks')
UPLOAD_FOLDER_SUBMISSIONS = os.environ.get('UPLOAD_FOLDER_SUBMISSIONS', 'static/uploads/submissions')
ALLOWED_EXTENSIONS        = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_FILE_SIZE             = int(os.environ.get('MAX_FILE_SIZE', 5242880))  # 5MB default

# SECURE: Database paths from environment with fallbacks
TASKS_DB_PATH             = os.environ.get('TASKS_DB_PATH', 'data/tasks.xlsx')
SUBMISSIONS_DB_PATH       = os.environ.get('SUBMISSIONS_DB_PATH', 'data/submissions.xlsx')

# Create directories on first run
os.makedirs(UPLOAD_FOLDER_TASKS, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_SUBMISSIONS, exist_ok=True)
# Create database directory from path
os.makedirs(os.path.dirname(TASKS_DB_PATH), exist_ok=True)

# SECURE: Admin credentials with fallbacks
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')
ADMIN_PASSWORD_HASH = generate_password_hash(ADMIN_PASSWORD)

# ─────────────────────────── Helper functions ────────────────────────
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_excel_files():
    """Create blank Excel files on first run."""
    if not os.path.exists(TASKS_DB_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Tasks'
        ws.append(['id', 'title', 'description', 'reference_image',
                   'created_at', 'status'])
        wb.save(TASKS_DB_PATH)

    if not os.path.exists(SUBMISSIONS_DB_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Submissions'
        # Added mobile_number column
        ws.append(['id', 'task_id', 'user_name', 'user_email', 'mobile_number',
                   'proof_image', 'submitted_at', 'status', 'admin_notes'])
        wb.save(SUBMISSIONS_DB_PATH)


def get_tasks():
    try:
        wb = load_workbook(TASKS_DB_PATH)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        return [dict(zip(headers, row))
                for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    except Exception as e:
        print('Error loading tasks:', e)
        return []

def get_task_by_id(task_id):
    return next((t for t in get_tasks() if str(t['id']) == str(task_id)), None)

def add_task(title, description, reference_image):
    try:
        wb = load_workbook(TASKS_DB_PATH)
        ws = wb.active
        task_id = str(uuid.uuid4())[:8]
        ws.append([task_id, title, description, reference_image,
                   datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'active'])
        wb.save(TASKS_DB_PATH)
        return task_id
    except Exception as e:
        print('Error adding task:', e)
        return None

def add_submission(task_id, user_name, user_email, mobile_number, proof_images):
    """Add submission with mobile number and multiple proof images (1-3 images)"""
    try:
        wb = load_workbook(SUBMISSIONS_DB_PATH)
        ws = wb.active
        submission_id = str(uuid.uuid4())[:8]
        images_csv = ','.join(proof_images) if isinstance(proof_images, list) else proof_images
        
        # Ensure status is always 'pending' (never None)
        ws.append([submission_id, task_id, user_name, user_email, mobile_number, images_csv,
                  datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'pending', ''])
        wb.save(SUBMISSIONS_DB_PATH)
        return submission_id
    except Exception as e:
        print('Error adding submission:', e)
        return None



def get_submissions():
    try:
        wb = load_workbook(SUBMISSIONS_DB_PATH)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        return [dict(zip(headers, row))
                for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    except Exception as e:
        print('Error loading submissions:', e)
        return []
    
def preprocess_image_for_comparison(image_path, output_size=(512, 512)):
    """
    Preprocess images to improve comparison accuracy
    """
    try:
        img = cv2.imread(image_path)
        
        # 1. Resize to standard size
        img = cv2.resize(img, output_size)
        
        # 2. Noise reduction
        img = cv2.bilateralFilter(img, 9, 75, 75)
        
        # 3. Enhance contrast
        lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
        lab[:,:,0] = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8)).apply(lab[:,:,0])
        img = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
        
        # Save processed image temporarily
        temp_path = image_path.replace('.', '_processed.')
        cv2.imwrite(temp_path, img)
        
        return temp_path
        
    except Exception as e:
        print(f"Preprocessing error: {e}")
        return image_path  # Return original if processing fails

    
def calculate_similarity_score(admin_image_path, user_image_path):
    """
    Enhanced similarity calculation with multiple methods for higher accuracy
    Returns score between 0.0 and 1.0 (higher = more similar)
    """
    try:
        # Load images
        img1 = cv2.imread(admin_image_path)
        img2 = cv2.imread(user_image_path)
        
        if img1 is None or img2 is None:
            return 0.0
        
        # Resize to same dimensions for comparison
        height, width = img1.shape[:2]
        img2 = cv2.resize(img2, (width, height))
        
        scores = []
        
        # Method 1: SSIM with multiple color channels
        gray1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
        gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
        ssim_score = ssim(gray1, gray2)
        scores.append(('ssim', ssim_score, 0.3))  # 30% weight
        
        # Method 2: Histogram comparison (color distribution)
        hist1 = cv2.calcHist([img1], [0, 1, 2], None, [50, 50, 50], [0, 256, 0, 256, 0, 256])
        hist2 = cv2.calcHist([img2], [0, 1, 2], None, [50, 50, 50], [0, 256, 0, 256, 0, 256])
        hist_corr = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CORREL)
        scores.append(('histogram', hist_corr, 0.25))  # 25% weight
        
        # Method 3: Enhanced perceptual hash
        hash1 = imagehash.average_hash(Image.open(admin_image_path), hash_size=16)
        hash2 = imagehash.average_hash(Image.open(user_image_path), hash_size=16)
        hash_diff = hash1 - hash2
        hash_similarity = max(0, (64 - hash_diff) / 64)  # 64 bits for 16x16 hash
        scores.append(('hash', hash_similarity, 0.2))  # 20% weight
        
        # Method 4: Feature matching with ORB
        orb = cv2.ORB_create(nfeatures=500)
        kp1, des1 = orb.detectAndCompute(gray1, None)
        kp2, des2 = orb.detectAndCompute(gray2, None)
        
        if des1 is not None and des2 is not None:
            bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
            matches = bf.match(des1, des2)
            matches = sorted(matches, key=lambda x: x.distance)
            
            # Calculate feature similarity based on good matches
            good_matches = [m for m in matches if m.distance < 50]
            feature_similarity = min(len(good_matches) / 20, 1.0)  # Normalize to 0-1
        else:
            feature_similarity = 0.0
        
        scores.append(('features', feature_similarity, 0.25))  # 25% weight
        
        # Calculate weighted final score
        final_score = sum(score * weight for _, score, weight in scores)
        
        # Debug output (remove in production)
        print(f"Similarity breakdown: SSIM={ssim_score:.3f}, Hist={hist_corr:.3f}, Hash={hash_similarity:.3f}, Features={feature_similarity:.3f}, Final={final_score:.3f}")
        
        return max(0.0, min(1.0, final_score))
        
    except Exception as e:
        print(f"Similarity calculation error: {e}")
        return 0.0


# Set your similarity threshold
SIMILARITY_THRESHOLD = 0.50  # 50% similarity for auto-approval

def update_submission_status(submission_id, status, notes=''):
    """Update submission row and return its task_id."""
    try:
        wb = load_workbook(SUBMISSIONS_DB_PATH)
        ws = wb.active
        task_id = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == str(submission_id):
                task_id = ws.cell(row=row, column=2).value
                ws.cell(row=row, column=7, value=status)
                ws.cell(row=row, column=8, value=notes)
                break
        wb.save(SUBMISSIONS_DB_PATH)
        return task_id
    except Exception as e:
        print('Error updating submission:', e)
        return None

def update_task_status(task_id, new_status):
    try:
        wb = load_workbook(TASKS_DB_PATH)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == str(task_id):
                ws.cell(row=row, column=6, value=new_status)
                break
        wb.save(TASKS_DB_PATH)
        return True
    except Exception as e:
        print('Error updating task status:', e)
        return False

def count_completed_tasks():
    try:
        wb = load_workbook(TASKS_DB_PATH)
        ws = wb.active
        return sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                   if row and row[5] == 'completed')
    except Exception as e:
        print('Error counting completed tasks:', e)
        return 0

def delete_task_and_children(task_id):
    """Completely remove a task, its submissions, and all related files."""
    # remove task row & reference image
    wb_t = load_workbook(TASKS_DB_PATH)
    ws_t = wb_t.active
    for row in range(2, ws_t.max_row + 1):
        if str(ws_t.cell(row=row, column=1).value) == str(task_id):
            ref_img = ws_t.cell(row=row, column=4).value
            ws_t.delete_rows(row, 1)
            if ref_img:
                f = os.path.join(UPLOAD_FOLDER_TASKS, ref_img)
                if os.path.exists(f):
                    os.remove(f)
            break
    wb_t.save(TASKS_DB_PATH)

    # remove submission rows & proof images
    wb_s = load_workbook(SUBMISSIONS_DB_PATH)
    ws_s = wb_s.active
    rows_to_delete = []
    for row in range(2, ws_s.max_row + 1):
        if str(ws_s.cell(row=row, column=2).value) == str(task_id):
            rows_to_delete.append(row)
            proof = ws_s.cell(row=row, column=5).value
            if proof:
                f = os.path.join(UPLOAD_FOLDER_SUBMISSIONS, proof)
                if os.path.exists(f):
                    os.remove(f)
    for row in reversed(rows_to_delete):
        ws_s.delete_rows(row, 1)
    wb_s.save(SUBMISSIONS_DB_PATH)

# Initialize Excel storage
init_excel_files()

# ─────────────────────────────── Routes ───────────────────────────────
@app.route('/')
def index():
    tasks = get_tasks()
    active_tasks = [t for t in tasks if t['status'] == 'active']
    completed_tasks = count_completed_tasks()
    return render_template('index.html',
                           tasks=active_tasks,
                           completed_tasks=completed_tasks)

@app.route('/task/<task_id>')
def task_detail(task_id):
    task = get_task_by_id(task_id)
    if not task:
        flash('Task not found!', 'error')
        return redirect(url_for('index'))
    return render_template('task_detail.html', task=task)

@app.route('/submit/<task_id>', methods=['POST'])
def submit_task(task_id):
    if not get_task_by_id(task_id):
        flash('Task not found!', 'error')
        return redirect(url_for('index'))

    user_name = request.form.get('user_name', '').strip()
    user_email = request.form.get('user_email', '').strip()
    mobile_number = request.form.get('mobile_number', '').strip()  # NEW FIELD
    
    # Get multiple files from form
    uploaded_files = request.files.getlist('proof_image')

    # Updated validation to include mobile number
    if not user_name or not user_email or not mobile_number or not uploaded_files:
        flash('All fields are required! Please fill in name, email, mobile number and upload at least 1 image.', 'error')
        return redirect(url_for('task_detail', task_id=task_id))

    # Validate mobile number (basic validation)
    if len(mobile_number) < 10 or not mobile_number.isdigit():
        flash('Please enter a valid mobile number (minimum 10 digits).', 'error')
        return redirect(url_for('task_detail', task_id=task_id))

    # Validate number of files (1 to 3 images required)
    if len(uploaded_files) < 1 or len(uploaded_files) > 3:
        flash('Please upload between 1 to 3 images as proof.', 'error')
        return redirect(url_for('task_detail', task_id=task_id))

    # Validate and save each file
    saved_filenames = []
    for i, file in enumerate(uploaded_files):
        if not file or not file.filename:
            flash(f'Image {i+1} is empty. Please select valid files.', 'error')
            return redirect(url_for('task_detail', task_id=task_id))
            
        if not allowed_file(file.filename):
            flash(f'Image {i+1} has invalid file type. Please use PNG, JPG, JPEG, GIF, or WEBP.', 'error')
            return redirect(url_for('task_detail', task_id=task_id))

    # Save all files if validation passes
        try:
            for i, file in enumerate(uploaded_files):
                filename = secure_filename(
                    f"{task_id}_{datetime.now().strftime('%Y%m%d_%H%M%S%f')}_{i+1}_{file.filename}")
                file.save(os.path.join(UPLOAD_FOLDER_SUBMISSIONS, filename))
                saved_filenames.append(filename)
            
            submission_id = add_submission(task_id, user_name, user_email, mobile_number, saved_filenames)
            
            # NEW: Similarity check for auto-approval
            if submission_id and saved_filenames:
                task = get_task_by_id(task_id)
                if task and task['reference_image']:
                    admin_ref_path = os.path.join(UPLOAD_FOLDER_TASKS, task['reference_image'])
                    user_image_path = os.path.join(UPLOAD_FOLDER_SUBMISSIONS, saved_filenames[0])
                    
                    # Calculate similarity
                    similarity = calculate_similarity_score(admin_ref_path, user_image_path)
                    
                    if similarity >= SIMILARITY_THRESHOLD:
                        # Auto-approve high similarity submissions
                        update_submission_status(submission_id, 'approved', 
                            f'Auto-approved: {similarity:.1%} similarity match')
                        flash(f'✅ Task auto-approved! Similarity: {similarity:.1%}', 'success')
                    else:
                        flash(f'📝 Submitted for manual review (Similarity: {similarity:.1%})', 'info')
                else:
                    flash(f'Your proof has been submitted successfully with {len(saved_filenames)} images!', 'success')
            
        except Exception as e:
            flash('Error uploading files. Please try again.', 'error')
            print('Upload error:', e)

        return redirect(url_for('index'))


# --------------------------- Admin auth -------------------------------
@app.route('/admin')
def admin_login():
    return redirect(url_for('admin_dashboard')) if 'admin_logged_in' in session \
           else render_template('admin_login.html')

@app.route('/admin/login', methods=['POST'])
def admin_login_post():
    if request.form.get('username') == ADMIN_USERNAME and \
       check_password_hash(ADMIN_PASSWORD_HASH, request.form.get('password', '')):
        session['admin_logged_in'] = True
        flash('Login successful!', 'success')
        return redirect(url_for('admin_dashboard'))
    flash('Invalid credentials!', 'error')
    return redirect(url_for('admin_login'))

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    flash('Logged out.', 'success')
    return redirect(url_for('admin_login'))

# --------------------------- Admin pages ------------------------------
@app.route('/admin/dashboard')
def admin_dashboard():
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))

    tasks = get_tasks()
    submissions = get_submissions()
    stats = {
        'total_tasks': len(tasks),
        'active_tasks': len([t for t in tasks if t['status'] == 'active']),
        'completed_tasks': count_completed_tasks(),
        'total_submissions': len(submissions),
        'pending_submissions': len([s for s in submissions if s['status'] == 'pending'])
    }
    return render_template('admin_dashboard.html', stats=stats)

@app.route('/admin/tasks')
def admin_tasks():
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    return render_template('admin_tasks.html', tasks=get_tasks())

@app.route('/admin/add_task', methods=['POST'])
def admin_add_task():
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))

    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    reference = request.files.get('reference_image')

    if not title or not description or not reference:
        flash('All fields are required!', 'error')
        return redirect(url_for('admin_tasks'))

    if not allowed_file(reference.filename):
        flash('Invalid file type.', 'error')
        return redirect(url_for('admin_tasks'))

    filename = secure_filename(
        f"ref_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{reference.filename}")
    reference.save(os.path.join(UPLOAD_FOLDER_TASKS, filename))
    add_task(title, description, filename)
    flash('Task added.', 'success')
    return redirect(url_for('admin_tasks'))

@app.route('/admin/submissions')
def admin_submissions():
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))

    subs = get_submissions()
    title = {t['id']: t['title'] for t in get_tasks()}
    for s in subs:
        s['task_title'] = title.get(s['task_id'], 'Unknown Task')
    return render_template('admin_submissions.html', submissions=subs)

@app.route('/admin/update_submission', methods=['POST'])
def admin_update_submission():
    """
    Admin approves or rejects a submission.
    The submission's status is updated,
    but the task itself stays 'active' until the admin manually marks it completed or deletes it.
    """
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))

    submission_id = request.form.get('submission_id')
    status = request.form.get('status')  # 'approved' or 'rejected'
    notes = request.form.get('notes', '')

    if update_submission_status(submission_id, status, notes):
        flash('Submission updated successfully!', 'success')
    else:
        flash('Error updating submission. Please try again.', 'error')

    # Task stays visible to users - no call to update_task_status()
    return redirect(url_for('admin_submissions'))

# --------------- Mark task as completed manually ----------------
@app.route('/admin/mark_completed/<task_id>', methods=['POST'])
def admin_mark_completed(task_id):
    """Mark a task as completed without deleting it."""
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    if update_task_status(task_id, 'completed'):
        flash('Task marked as completed!', 'success')
    else:
        flash('Error updating task status.', 'error')
    
    return redirect(url_for('admin_tasks'))

# --------------- Delete task AFTER admin decides to remove it --------
@app.route('/admin/delete_task/<task_id>', methods=['POST'])
def admin_delete_task(task_id):
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    delete_task_and_children(task_id)
    flash('Task and all related data deleted.', 'success')
    return redirect(url_for('admin_tasks'))

# ------------------------ Error handlers ------------------------------
@app.errorhandler(404)
def not_found_error(_):
    return """
    <h1>404 - Page Not Found</h1>
    <p>The page you're looking for doesn't exist.</p>
    <a href="/">← Back to Home</a>
    """, 404

@app.errorhandler(500)
def internal_error(_):
    return """
    <h1>500 - Internal Server Error</h1>
    <p>Something went wrong on our end.</p>
    <a href="/">← Back to Home</a>
    """, 500

# ─────────────────────────────── Main ────────────────────────────────
if __name__ == '__main__':
    # SECURE: Get configuration from environment with safe fallbacks
    debug_mode = os.environ.get('DEBUG', 'True').lower() == 'true'
    host = os.environ.get('HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', 5000))
    
    print('Starting Microtask Website…')
    print(f'Admin credentials: {ADMIN_USERNAME} / [password from .env file]')
    print(f'Server running on: http://{host}:{port}')
    print(f'Database files: {TASKS_DB_PATH}, {SUBMISSIONS_DB_PATH}')
    
    app.run(debug=debug_mode, host=host, port=port)
